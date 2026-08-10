#!/usr/bin/env python3
"""Validate Avito ImageUrls cells: cloud-api format, count, varied firsts.

Usage:
  python validate_image_urls.py --xlsx path/to.xlsx --sheet "Кухонные гарнитуры-Кухни" --expect 4
  python validate_image_urls.py --xlsx path/to.xlsx --sheet "..." --expect 10 --root /AUTOZA/Foto_mebel_1
"""

from __future__ import annotations

import argparse
import re
from collections import Counter

import openpyxl

PREFIX = "https://cloud-api.yandex.net/v1/disk/resources/download?path="
CYRILLIC = re.compile(r"[А-Яа-яЁё]")


def headers_map(ws):
    # row 2 = human headers in Avito xlsx; also accept row 1 Id-style templates
    mapping = {}
    for c in range(1, ws.max_column + 1):
        for r in (1, 2, 3, 4):
            v = ws.cell(r, c).value
            if not v:
                continue
            key = str(v).strip()
            if key and key not in mapping:
                mapping[key] = c
    return mapping


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--xlsx", required=True)
    p.add_argument("--sheet", required=True)
    p.add_argument("--expect", type=int, default=4, help="Expected URL count per row (sets pipeline=4, full=10)")
    p.add_argument(
        "--root",
        default="/AUTOZA/Foto_mebel_1",
        help="Required path prefix after ?path= (empty string to skip check)",
    )
    p.add_argument("--check-names-empty", action="store_true", help="Fail if ImageNames filled")
    args = p.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    ws = wb[args.sheet]
    h = headers_map(ws)
    url_col = h.get("Ссылки на фото") or h.get("ImageUrls") or h.get("Image Urls")
    # Avito template often has ImageUrls in row1 technical / row2 Russian
    if not url_col:
        for k, c in h.items():
            if "ImageUrl" in k.replace(" ", "") or "Ссылки на фото" in k:
                url_col = c
                break
    if not url_col:
        raise SystemExit(f"No ImageUrls column; headers={list(h)[:20]}")

    names_col = None
    if args.check_names_empty:
        names_col = h.get("Названия фото") or h.get("ImageNames")

    firsts = Counter()
    bad = 0
    rows = 0
    for r in range(5, ws.max_row + 1):
        raw = ws.cell(r, url_col).value
        if not raw:
            continue
        parts = [x.strip() for x in str(raw).split("|") if x.strip()]
        rows += 1
        if len(parts) != args.expect:
            bad += 1
            print(f"R{r}: count={len(parts)} expected={args.expect}")
        for u in parts:
            if not u.startswith(PREFIX):
                bad += 1
                print(f"R{r}: bad prefix {u[:80]}")
            if CYRILLIC.search(u):
                bad += 1
                print(f"R{r}: cyrillic in path")
            if args.root and PREFIX in u:
                path = u.split("path=", 1)[-1]
                if not path.startswith(args.root.rstrip("/") + "/") and path != args.root:
                    bad += 1
                    print(f"R{r}: path root mismatch {path[:80]}")
        if names_col and ws.cell(r, names_col).value:
            bad += 1
            print(f"R{r}: ImageNames should be empty")
        if parts:
            firsts[parts[0].rsplit("/", 1)[-1]] += 1

    print(f"rows={rows} bad={bad} unique_firsts={len(firsts)} expect={args.expect} root={args.root!r}")
    if firsts:
        print("top firsts:", firsts.most_common(5))
        if len(firsts) == 1 and rows > 3:
            print("FAIL: all ads share the same first photo")
            raise SystemExit(1)
    wb.close()
    if bad:
        raise SystemExit(1)
    print("OK")


if __name__ == "__main__":
    main()
